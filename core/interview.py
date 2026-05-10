"""인터뷰지 로드·진행·저장 모듈.

기존 인터뷰지 (Part 1 정량지표 + Part 2 심층 인터뷰)를 구조화된 질문 리스트로 변환.
각 질문에 고유 ID와 정보 태그를 부여하여 양식 중립적 매핑이 가능하게 한다.

Excel 구조 (심층 인터뷰 시트):
  col[0]: (empty / "Part.2")
  col[1]: 카테고리 번호 ("1"/"2"/"3"/"4") 또는 empty
  col[2]: 카테고리명("Problem") / 섹션명("1-1. 개발동기") / 분기("공통"/"제조"/"IT")
  col[3]: "Q" / "A"
  col[4]: 질문 본문
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


CATEGORY_NAMES = {"Problem", "Solution", "Business Model", "Team & Infrastructure"}
BRANCH_NAMES = {"공통", "제조", "IT"}
SECTION_RE = re.compile(r"^\d+-\d+\.?\s")


@dataclass
class Question:
    qid: str
    section: str
    category: str
    text: str
    branch: Optional[str] = None
    tags: list[str] = field(default_factory=list)


@dataclass
class Answer:
    qid: str
    text: str
    updated_at: Optional[str] = None


@dataclass
class Session:
    session_id: str
    program_code: str
    answers: dict[str, Answer] = field(default_factory=dict)
    company_context: Optional[dict] = None  # 인터뷰 완료 후 1회 전처리된 8개 정제 항목

    def to_json(self) -> str:
        return json.dumps(
            {
                "session_id": self.session_id,
                "program_code": self.program_code,
                "answers": {qid: asdict(a) for qid, a in self.answers.items()},
                "company_context": self.company_context,
            },
            ensure_ascii=False,
            indent=2,
        )

    @classmethod
    def from_json(cls, text: str) -> "Session":
        data = json.loads(text)
        return cls(
            session_id=data["session_id"],
            program_code=data["program_code"],
            answers={qid: Answer(**a) for qid, a in data["answers"].items()},
            company_context=data.get("company_context"),
        )


_initial_questions_cache: dict[str, list[Question]] = {}


def load_initial_questions(json_path: str | Path) -> list[Question]:
    """초기 인터뷰용 10개 통합 질문을 JSON에서 로드."""
    key = str(Path(json_path).resolve())
    if key in _initial_questions_cache:
        return _initial_questions_cache[key]
    path = Path(json_path)
    data = json.loads(path.read_text(encoding="utf-8"))
    questions = [
        Question(
            qid=item["qid"],
            section=item["section"],
            category=item["category"],
            text=item["text"],
            branch=item.get("branch", "공통"),
            tags=item.get("tags", []),
        )
        for item in data
    ]
    _initial_questions_cache[key] = questions
    return questions


_followup_questions_cache: dict[str, list[Question]] = {}


def load_followup_questions(json_path: str | Path) -> list[Question]:
    """초안 생성 후 LLM이 메모 생성 시 참고할 60개 후속 질문을 JSON에서 로드."""
    key = str(Path(json_path).resolve())
    if key in _followup_questions_cache:
        return _followup_questions_cache[key]
    path = Path(json_path)
    data = json.loads(path.read_text(encoding="utf-8"))
    questions = [
        Question(
            qid=item["qid"],
            section=item["section"],
            category=item.get("category", ""),
            text=item["text"],
            branch=item.get("branch", "공통"),
            tags=item.get("tags", []),
        )
        for item in data
    ]
    _followup_questions_cache[key] = questions
    return questions


def load_questions(xlsx_path: str | Path) -> list[Question]:
    wb = load_workbook(xlsx_path, data_only=True)
    questions: list[Question] = []

    if "심층 인터뷰" not in wb.sheetnames:
        return questions

    ws = wb["심층 인터뷰"]
    current_category: Optional[str] = None
    current_section: Optional[str] = None
    current_branch: str = "공통"
    section_counter: dict[tuple[str, str], int] = {}

    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        # Pad to length 9
        while len(cells) < 9:
            cells.append("")

        c1, c2, c3, c4 = cells[1], cells[2], cells[3], cells[4]

        # 카테고리 헤더: col1=숫자, col2=카테고리명
        if c1 in {"1", "2", "3", "4"} and c2 in CATEGORY_NAMES:
            current_category = c2
            current_section = None
            current_branch = "공통"
            continue

        # 섹션 헤더: col2가 "1-1.", "2-2.", ... 으로 시작
        if c2 and SECTION_RE.match(c2):
            current_section = c2
            current_branch = "공통"
            continue

        # 분기 표시: col2가 "공통"/"제조"/"IT"
        if c2 in BRANCH_NAMES:
            current_branch = c2
            # 같은 행에 Q가 함께 오는 경우도 있음 → 이어서 처리

        # 질문 행: col3="Q", col4=텍스트
        if c3 == "Q" and c4:
            if current_category is None or current_section is None:
                continue
            sec_id = current_section.split(".")[0].strip()
            key = (sec_id, current_branch)
            section_counter[key] = section_counter.get(key, 0) + 1
            qnum = section_counter[key]
            qid = (
                f"{sec_id}-Q{qnum}"
                if current_branch == "공통"
                else f"{sec_id}-{current_branch}-Q{qnum}"
            )
            questions.append(
                Question(
                    qid=qid,
                    section=current_section,
                    category=current_category,
                    text=c4,
                    branch=current_branch,
                    tags=_guess_tags(current_section, c4),
                )
            )

    return questions


def _guess_tags(section: str, q_text: str) -> list[str]:
    tags: list[str] = []
    mapping = [
        (["개발동기", "개발목적"], "개발동기"),
        (["시장분석", "시장"], "시장분석"),
        (["고객"], "고객"),
        (["경쟁"], "경쟁사"),
        (["비즈니스 모델"], "BM"),
        (["시장 진입", "사업화 전략"], "사업화전략"),
        (["일정", "자금"], "일정자금"),
        (["기업구성", "역량"], "팀역량"),
        (["ESG"], "ESG"),
    ]
    for keywords, tag in mapping:
        if any(k in section for k in keywords):
            tags.append(tag)

    text_rules = [
        (["특허", "상표"], "IP"),
        (["투자"], "투자"),
        (["매출", "수익"], "재무"),
        (["해외", "수출"], "해외"),
        (["대표"], "대표자"),
    ]
    for keywords, tag in text_rules:
        if any(k in q_text for k in keywords):
            tags.append(tag)

    return list(dict.fromkeys(tags))


def save_session(session: Session, dir_path: str | Path = "data/sessions") -> Path:
    path = Path(dir_path)
    path.mkdir(parents=True, exist_ok=True)
    file_path = path / f"{session.session_id}.json"
    file_path.write_text(session.to_json(), encoding="utf-8")
    return file_path


def load_session(session_id: str, dir_path: str | Path = "data/sessions") -> Optional[Session]:
    file_path = Path(dir_path) / f"{session_id}.json"
    if not file_path.exists():
        return None
    return Session.from_json(file_path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    path = Path(__file__).resolve().parent.parent / "data" / "interview" / "interview_v1.xlsx"
    qs = load_questions(path)
    print(f"Loaded {len(qs)} questions")
    by_section: dict[str, int] = {}
    for q in qs:
        by_section[q.section] = by_section.get(q.section, 0) + 1
    for sec, cnt in by_section.items():
        print(f"  {sec}: {cnt}")
    print("\n--- first 3 ---")
    for q in qs[:3]:
        print(f"  [{q.qid}] ({q.branch}) {q.text[:60]}  tags={q.tags}")
    print("\n--- last 3 ---")
    for q in qs[-3:]:
        print(f"  [{q.qid}] ({q.branch}) {q.text[:60]}  tags={q.tags}")
